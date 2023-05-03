import asyncio
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor
from multiprocessing import Process

from PyQt5.QtCore import QThreadPool, QRunnable

from scrapy_thread import ScrapyThread

if sys.version_info < (3, 9):
    asyncio.run = asyncio.run_coroutine_threadsafe
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QVBoxLayout, QHBoxLayout, QPushButton, \
    QSystemTrayIcon, QMenu, QMessageBox
from PyQt5.QtGui import QIcon
from openpyxl import Workbook, load_workbook
from scrapy import cmdline

from settings import *


class ScrapyTool(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Scrapy Tool')
        self.setWindowIcon(QIcon('zone.ico'))

        self.link_edit = QLineEdit()
        self.title_edit = QLineEdit()
        self.content_edit = QLineEdit()

        self.ok_button = QPushButton('确定')
        self.cancel_button = QPushButton('取消')

        link_label = QLabel('网页链接：')
        title_label = QLabel('标题CSS：')
        content_label = QLabel('内容CSS：')

        link_layout = QHBoxLayout()
        link_layout.addWidget(link_label)
        link_layout.addWidget(self.link_edit)

        title_layout = QHBoxLayout()
        title_layout.addWidget(title_label)
        title_layout.addWidget(self.title_edit)

        content_layout = QHBoxLayout()
        content_layout.addWidget(content_label)
        content_layout.addWidget(self.content_edit)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)

        main_layout = QVBoxLayout()
        main_layout.addLayout(link_layout)
        main_layout.addLayout(title_layout)
        main_layout.addLayout(content_layout)
        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

        self.ok_button.clicked.connect(self.save_data)
        self.cancel_button.clicked.connect(self.on_close)
        self.create_tray_icon()

    def save_data(self):
        file_name = EXCEL_FILE
        url = self.link_edit.text()
        title = self.title_edit.text()
        content = self.content_edit.text()

        if not url or not title or not content:
            QMessageBox.warning(self, "添加失败", "必输信息不能为空")
            return

        try:
            if not os.path.exists(APP_DIR):
                os.makedirs(APP_DIR)
            if not os.path.exists(file_name):
                wb = Workbook()
                ws = wb.active
                ws["A1"] = "网页链接"
                ws["B1"] = "标题css"
                ws["C1"] = "内容css"
            else:
                wb = load_workbook(file_name)
                ws = wb.active

            ws["A2"] = url
            ws["B2"] = title
            ws["C2"] = content
            wb.save(file_name)

            self.link_edit.clear()
            self.title_edit.clear()
            self.content_edit.clear()
            # 创建一个新进程并启动它
            self.start_scrapy()
        except Exception as e:
            QMessageBox.warning(self, "添加失败", f"出现错误: {e}")

    def on_close(self):
        self.quit_application()

    def create_tray_icon(self):
        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(QIcon('zone.ico'))  # 将路径替换为您的图标文件路径
        self.tray_icon.activated.connect(self.tray_icon_clicked)

        tray_menu = QMenu()
        exit_action = tray_menu.addAction("退出")
        exit_action.triggered.connect(self.quit_application)

        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.show()

    def tray_icon_clicked(self, reason):
        if reason == QSystemTrayIcon.Trigger:
            self.show()

    def closeEvent(self, a0) -> None:
        self.quit_application()

    def quit_application(self):
        self.tray_icon.hide()
        QApplication.quit()

    def start_scrapy(self):
        subprocess.run(["scrapy", "crawl", "common"])

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    scrapy_tool = ScrapyTool()
    scrapy_tool.show()
    sys.exit(app.exec_())
