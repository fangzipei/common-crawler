import scrapy
import logging
from common.items import ProductItem
from scrapy import cmdline
import openpyxl
from settings import *
from datetime import datetime

log = logging.getLogger("common")
today = datetime.today().strftime('%Y-%m-%d')


class CommonSpider(scrapy.Spider):
    name = "common"

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        workbook = openpyxl.load_workbook(EXCEL_FILE)

        ws = workbook.active
        self.start_urls = [str(ws["A2"].value)]
        self.title = str(ws["B2"].value)
        self.content = str(ws["C2"].value)

    def extract_item(self, response):
        title = response.css(self.title).get()  # 获取标题
        content = response.css(self.content).get()  # 获取内容
        item = ProductItem()
        item['title'] = title
        item['content'] = content
        return item

    def parse(self, response, **kwargs):
        try:
            item = self.extract_item(response)
            yield item
        except Exception as e:
            log.error(e)
            log.error("解析分页信息出错，url：%s", response.url)


if __name__ == "__main__":
    cmdline.execute('scrapy crawl common'.split())
